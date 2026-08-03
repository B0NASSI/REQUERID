# -*- coding: utf-8 -*-
"""Leitura de planilhas .xlsx ou .csv para o modo lote, sem depender de pandas.

Suporta dois formatos de planilha:
  1. Esquema próprio do projeto: colunas nomeadas empresa/segurado/cpf/nit/
     especie/nb (mais sexo e data, opcionais) - ver dados/exemplo.xlsx.
  2. Planilha "SOLICITAÇÕES GERID" usada no dia a dia com cada cliente:
     colunas Empresa / Número do Benefício / Espécie / NIT do Empregado
     (dígitos puros) / CPF do Empregado (ou "NIT do Empregado" de novo, nas
     planilhas antigas ainda não corrigidas - erro de digitação na planilha
     original, identificado pelo dígito verificador de cada um) / Segurados
     - ver dados/Exemplo.xlsx. Aceita os dois cabeçalhos (antigo
     com "NIT do Empregado" repetido, ou já corrigido com "CPF do
     Empregado"); por isso esse formato é lido por posição, não por nome.
     Aceita também colunas opcionais "Item" e "CNPJ" inseridas em qualquer
     posição (ex.: "Item" no início, "CNPJ" logo depois de "Empresa"); as
     demais 6 colunas-base continuam sendo localizadas pela ordem relativa
     entre si. Tolera ainda uma linha de título acima do cabeçalho de verdade
     (ex.: "PEDIDO DE CÓPIA DE PROCESSO - ..."), procurando o cabeçalho nas
     primeiras linhas da planilha em vez de assumir que é sempre a primeira.
"""

import csv
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

from core import CAMPOS_OBRIGATORIOS, ErroValidacao

CHAVES_BASE_SOLICITACOES_GERID = ["empresa", "nb", "especie", "nit", "cpf", "segurado"]

# Dois cabeçalhos aceitos para a 6ª/5ª coluna-base (CPF do segurado): a
# planilha original tinha um erro de digitação e repetia "NIT do Empregado"
# duas vezes (a 2ª era na verdade o CPF); algumas planilhas já foram
# corrigidas para "CPF do Empregado". Aceita os dois.
CABECALHOS_SOLICITACOES_GERID_BASE = [
    ["empresa", "numero do beneficio", "especie", "nit do empregado", "nit do empregado", "segurados"],
    ["empresa", "numero do beneficio", "especie", "nit do empregado", "cpf do empregado", "segurados"],
]


def _sem_acento(texto: str) -> str:
    norm = unicodedata.normalize("NFKD", str(texto))
    return "".join(c for c in norm if not unicodedata.combining(c))


def _normalizar_cabecalho(valor) -> str:
    return _sem_acento(str(valor)).strip().lower() if valor is not None else ""


def _linha_vazia(valores) -> bool:
    return all(v is None or str(v).strip() == "" for v in valores)


def _detectar_solicitacoes_gerid(cabecalho_bruto: list) -> dict | None:
    """Reconhece a planilha "SOLICITAÇÕES GERID" pelas 6 colunas-base, na
    ordem em que aparecem - ignorando colunas "cnpj" e "item" opcionais, que
    podem estar em qualquer posição (ex.: "Item" no início, "CNPJ" depois de
    "Empresa"). Retorna um dict com o índice de cada campo nas colunas
    originais, ou None se o cabeçalho não corresponder a esse formato."""
    normalizados = [_normalizar_cabecalho(c) for c in cabecalho_bruto]
    indices_cnpj = [i for i, c in enumerate(normalizados) if c.startswith("cnpj")]
    indices_item = [i for i, c in enumerate(normalizados) if c == "item"]
    posicoes_base = [i for i, c in enumerate(normalizados) if not c.startswith("cnpj") and c != "item"]
    candidata = [normalizados[i] for i in posicoes_base[:6]]

    if candidata not in CABECALHOS_SOLICITACOES_GERID_BASE:
        return None

    posicoes = dict(zip(CHAVES_BASE_SOLICITACOES_GERID, posicoes_base[:6]))
    posicoes["cnpj"] = indices_cnpj[0] if indices_cnpj else None
    posicoes["item"] = indices_item[0] if indices_item else None
    return posicoes


def _adaptador_solicitacoes_gerid(posicoes: dict):
    def adaptar(valores) -> dict:
        linha = {chave: valores[indice] for chave, indice in posicoes.items() if indice is not None}
        return linha

    return adaptar


LINHAS_BUSCA_CABECALHO = 5


def _localizar_cabecalho_solicitacoes(linhas: list) -> tuple[int, dict] | tuple[None, None]:
    """Procura o cabeçalho da planilha "SOLICITAÇÕES GERID" nas primeiras
    linhas (tolerando uma linha de título acima dele). Devolve (índice da
    linha do cabeçalho, posições) ou (None, None) se não encontrar."""
    for indice, linha in enumerate(linhas[:LINHAS_BUSCA_CABECALHO]):
        posicoes = _detectar_solicitacoes_gerid(list(linha))
        if posicoes is not None:
            return indice, posicoes
    return None, None


def _processar_linhas(todas_linhas: list) -> list[dict]:
    if not todas_linhas:
        return []

    indice_cabecalho, posicoes = _localizar_cabecalho_solicitacoes(todas_linhas)
    if posicoes is not None:
        adaptar = _adaptador_solicitacoes_gerid(posicoes)
        linhas_dados = todas_linhas[indice_cabecalho + 1 :]
    else:
        cabecalho = [_normalizar_cabecalho(v) for v in todas_linhas[0]]
        adaptar = lambda valores: dict(zip(cabecalho, valores))  # noqa: E731
        linhas_dados = todas_linhas[1:]

    registros = []
    for valores in linhas_dados:
        registro = adaptar(valores)
        # Ignora a coluna "item" (numeração sequencial pré-preenchida na
        # planilha-modelo) ao decidir se a linha está vazia - senão uma
        # linha só com o número do item (sem nenhum dado real) não seria
        # descartada como deveria.
        significativos = [v for k, v in registro.items() if k != "item"]
        if _linha_vazia(significativos):
            continue
        registros.append(registro)
    return registros


def ler_xlsx(caminho: Path) -> list[dict]:
    wb = load_workbook(str(caminho), data_only=True)
    planilha = wb.active
    return _processar_linhas(list(planilha.iter_rows(values_only=True)))


def ler_csv(caminho: Path) -> list[dict]:
    delimitador = ";" if _detectar_delimitador(caminho) else ","
    with open(caminho, newline="", encoding="utf-8-sig") as f:
        leitor = csv.reader(f, delimiter=delimitador)
        return _processar_linhas(list(leitor))


def _detectar_delimitador(caminho: Path) -> bool:
    with open(caminho, encoding="utf-8-sig") as f:
        primeira_linha = f.readline()
    return primeira_linha.count(";") > primeira_linha.count(",")


def ler_planilha(caminho: Path) -> list[dict]:
    if not caminho.exists():
        raise ErroValidacao(f"Planilha não encontrada: {caminho}")

    sufixo = caminho.suffix.lower()
    if sufixo == ".xlsx":
        registros = ler_xlsx(caminho)
    elif sufixo == ".csv":
        registros = ler_csv(caminho)
    else:
        raise ErroValidacao(f"Formato de planilha não suportado: {sufixo} (use .xlsx ou .csv)")

    if not registros:
        raise ErroValidacao(f"Planilha vazia ou sem dados: {caminho}")

    colunas_presentes = set(registros[0].keys())
    faltando = [c for c in CAMPOS_OBRIGATORIOS if c not in colunas_presentes]
    if faltando:
        raise ErroValidacao(
            f"Planilha sem a(s) coluna(s) obrigatória(s): {', '.join(faltando)}. "
            f"Colunas encontradas: {', '.join(sorted(colunas_presentes)) or '(nenhuma)'}"
        )

    return registros
